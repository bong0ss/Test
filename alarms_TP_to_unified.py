import re

from openpyxl import load_workbook

# ---- CONFIG ----
INPUT_XLSX = "HMIAlarms.xlsx"
OUTPUT_XLSX = "alarms_export_unified_fixed.xlsx"

TAGS_TO_FIX = {
    "ALARMS_DF2101",
    "ALARMS_RW",
    "ALARMS_WS1",
    "ALARMS_WS2",
    "ALARMS_WS3",
    "ALARMS_WS4",
    "ALARMS_WS5",
    "ALARMS_WS6",
    "ALARMS_WS7",
    "ALARMS_WS8",
    "ALARMS_WS9",
    "ALARMS_WS10",
    "ALARMS_WS11",
    "ALARMS_WS12",
    "ALARMS_WS13",
    "ALARMS_WS14",
    "ALARMS_WS15",
    "ALARMS_WS16",
    "ALARMS_WS17",
    "ALARMS_WS18",
    "ALARMS_WS19",
    "ALARMS_WS20",
    "ALARMS_WS21",
    "ALARMS_WS22",
    "ALARMS_WS23",
    "ALARMS_WS24",
    "ALARMS_BUFFER_COM_P&P_1_ANDON",
    "DIST_HANDLING_FAULTS",
    "ALARMS_CFL",
    "ALARMS_GENERAL",
    "ALARMS_BCR",
    "ALARMS_PROFINET",
}


def norm(s):
    return re.sub(r"\s+", " ", str(s).strip().lower())


def extract_alarm_parameter_1(fieldinfo):
    """
    Extracts Tag from <ref id = 0 ... Tag = XXX; ...>
    """
    if not isinstance(fieldinfo, str):
        return None

    m = re.search(
        r"id\s*=\s*0.*?Tag\s*=\s*([^;>\s]+)", fieldinfo, flags=re.IGNORECASE | re.DOTALL
    )
    if m:
        return m.group(1).strip()
    return None


def extract_alarm_parameters(fieldinfo):
    """
    Extracts all Tags from <ref ... Tag = XXX; ...> blocks,
    ordered by their id (0, 1, 2, ...).
    """
    if not isinstance(fieldinfo, str):
        return []

    refs = re.findall(
        r"<ref\b.*?>",
        fieldinfo,
        flags=re.IGNORECASE | re.DOTALL,
    )

    def get_id(ref):
        m = re.search(r"id\s*=\s*(\d+)", ref, flags=re.IGNORECASE)
        return int(m.group(1)) if m else None

    refs_with_id = sorted(
        ((get_id(ref), ref) for ref in refs),
        key=lambda x: x[0] if x[0] is not None else 9999,
    )

    tags = []
    for _, ref in refs_with_id:
        m = re.search(
            r"Tag\s*=\s*([^;>\s]+)",
            ref,
            flags=re.IGNORECASE,
        )
        if m:
            tags.append(m.group(1).strip())

    return tags


def to_int(v):
    if isinstance(v, int):
        return v
    if isinstance(v, float) and v.is_integer():
        return int(v)
    if isinstance(v, str) and re.fullmatch(r"-?\d+", v.strip()):
        return int(v.strip())
    return None


def strip_index(tag):
    return re.sub(r"\[\d+\]\s*$", "", tag.strip())


def convert_fieldinfo(fieldinfo):
    if not isinstance(fieldinfo, str):
        return fieldinfo

    # Counter used to assign sequential Parameter numbers (1, 2, 3, ...)
    # based on appearance order of <ref ...> blocks, independent of id.
    param_counter = {"n": 0}

    def repl(m):
        ref = m.group(0)

        is_common = "type = CommonTextList" in ref

        param_counter["n"] += 1
        param_no = param_counter["n"]

        if ("type = AlarmTag" in ref) or is_common:
            ref = re.sub(
                r"type\s*=\s*(AlarmTag|CommonTextList)",
                f"type = AlarmParameterWithOrWithoutCommonTextList; "
                f"Parameter = Parameter {param_no}",
                ref,
                flags=re.IGNORECASE,
            )
        else:
            return ref

        extras = [
            "Precision = 0",
            "Alignment = Right",
            "ZeroPadding = False",
        ]

        if is_common and "DisplayType" not in ref:
            # Insert DisplayType = Textlist before Length if possible,
            # otherwise just append at the end together with other extras.
            if "Length" in ref:
                ref = re.sub(
                    r"(Length\s*=\s*[^;>]+;?)",
                    r"DisplayType = Textlist; \1",
                    ref,
                    flags=re.IGNORECASE,
                )
            else:
                extras.append("DisplayType = Textlist")

        for ex in extras:
            if ex not in ref:
                ref = ref.rstrip(">;") + f"; {ex};>"

        return ref

    return re.sub(r"<ref\b.*?>", repl, fieldinfo, flags=re.DOTALL)


def main():
    wb = load_workbook(INPUT_XLSX)
    ws = wb["DiscreteAlarms"]

    # Map headers
    headers = {}
    for c in range(1, ws.max_column + 1):
        v = ws.cell(1, c).value
        if v:
            headers[norm(v)] = c

    tag_col = headers["trigger tag"]
    bit_col = headers["trigger bit"]
    fieldinfo_col = headers["fieldinfo [alarm text]"]

    # ---- Rename headers ----
    ws.cell(1, 3).value = "Alarm text [en-US], Alarm text 1"
    ws.cell(1, 4).value = "FieldInfo [Alarm text 1]"

    # ---- Add "Alarm parameter" columns ----
    ws.insert_cols(ws.max_column + 1)
    param1_col = ws.max_column
    ws.cell(1, param1_col).value = "Alarm parameter 1"

    param2_col = param1_col + 1
    param3_col = param1_col + 2
    param4_col = param1_col + 3
    ws.cell(1, param2_col).value = "Alarm parameter 2"
    ws.cell(1, param3_col).value = "Alarm parameter 3"
    ws.cell(1, param4_col).value = "Alarm parameter 4"

    # ---- Process rows ----
    for r in range(2, ws.max_row + 1):
        # Trigger fix
        tag = ws.cell(r, tag_col).value
        bit = to_int(ws.cell(r, bit_col).value)

        if isinstance(tag, str) and bit is not None:
            base = strip_index(tag)
            if base in TAGS_TO_FIX:
                elem = bit // 16
                within = bit % 16
                fixed_bit = within ^ 8

                ws.cell(r, tag_col).value = f"{base}[{elem}]"
                ws.cell(r, bit_col).value = fixed_bit

        # FieldInfo conversion
        fi = ws.cell(r, fieldinfo_col).value
        ws.cell(r, fieldinfo_col).value = convert_fieldinfo(fi)

        # ---- Alarm parameter extraction ----
        fi = ws.cell(r, fieldinfo_col).value
        params = extract_alarm_parameters(fi)

        # Default values when parameter is missing
        ws.cell(r, param1_col).value = params[0] if len(params) >= 1 else "<No value>"
        ws.cell(r, param2_col).value = params[1] if len(params) >= 2 else "<No value>"
        ws.cell(r, param3_col).value = params[2] if len(params) >= 3 else "<No value>"
        ws.cell(r, param4_col).value = params[3] if len(params) >= 4 else "<No value>"

    wb.save(OUTPUT_XLSX)
    print(f"Done → {OUTPUT_XLSX}")


if __name__ == "__main__":
    main()
