import json
import os
import shutil
import time

import openpyxl
import pandas
from celery import shared_task
from celery_progress.backend import ProgressRecorder
from django.urls import reverse
from django.utils.safestring import mark_safe

from .tools.alarms_TP_to_unified import (
    convert_fieldinfo,
    extract_alarm_parameters,
    norm,
    strip_index,
    to_int,
)
from .tools.proface_adress_translator import (
    change_start_val,
    dbb_to_ls_word,
    dbx_to_ls,
    ls_bit_to_dbx,
    ls_to_dbb_bytes,
)
from .utility import custom_data


@shared_task(name="members.tasks.timer", bind=True)
def timer(self, time_left, user_id=None):
    progress = ProgressRecorder(self)
    custom_data(
        data=f"Task {self.request.id} started at {time.ctime()}, await {time_left} seconds! Predicted to finish at {time.ctime(time.time() + time_left)}"
    )
    for i in range(time_left):
        time.sleep(1)
        custom_data(data=f"Task {self.request.id}: {i}s.")
        progress.set_progress(i + 1, time_left, description="Processing...")
    custom_data(data=f"Task {self.request.id} finished at {time.ctime()}")
    return "Finished!"


@shared_task(bind=True)
def alarms_tp_uni(self, input_xlsx, output_xlsx, input_txt, user_id):
    custom_data(data=f"Task {self.request.id} started at {time.ctime()}.")
    custom_data(data=f"Task {self.request.id}: Loading workbook.")
    wb = openpyxl.load_workbook(input_xlsx)
    ws = wb["DiscreteAlarms"]
    og_output_xlsx = output_xlsx
    output_xlsx = str(user_id) + "_" + str(int(time.time())) + "_" + output_xlsx
    # Map headers
    custom_data(data=f"Task {self.request.id}: Mapping Excel Files.")
    headers = {}
    for c in range(1, ws.max_column + 1):
        v = ws.cell(1, c).value
        if v:
            headers[norm(v)] = c
    tag_col = headers["trigger tag"]
    bit_col = headers["trigger bit"]
    fieldinfo_col = headers["fieldinfo [alarm text]"]

    # ---- Rename headers ----
    custom_data(data=f"Task {self.request.id}: Renaming headers.")
    ws.cell(1, 3).value = "Alarm text [en-US], Alarm text 1"
    ws.cell(1, 4).value = "FieldInfo [Alarm text 1]"

    # ---- Add "Alarm parameter" columns ----
    custom_data(data=f"Task {self.request.id}: Adding Alarm parameters.")
    ws.insert_cols(ws.max_column + 1)
    param1_col = ws.max_column
    ws.cell(1, param1_col).value = "Alarm parameter 1"
    param2_col = param1_col + 1
    param3_col = param1_col + 2
    param4_col = param1_col + 3
    ws.cell(1, param2_col).value = "Alarm parameter 2"
    ws.cell(1, param3_col).value = "Alarm parameter 3"
    ws.cell(1, param4_col).value = "Alarm parameter 4"
    # ---- Process rows ----
    custom_data(data=f"Task {self.request.id}: Processing rows.")
    for r in range(2, ws.max_row + 1):
        # Trigger fix
        tag = ws.cell(r, tag_col).value
        bit = to_int(ws.cell(r, bit_col).value)

        if isinstance(tag, str) and bit is not None:
            base = strip_index(tag)
            if base in input_txt:
                elem = bit // 16
                within = bit % 16
                fixed_bit = within ^ 8

                ws.cell(r, tag_col).value = f"{base}[{elem}]"
                ws.cell(r, bit_col).value = fixed_bit

        # FieldInfo conversion
        fi = ws.cell(r, fieldinfo_col).value
        ws.cell(r, fieldinfo_col).value = convert_fieldinfo(fi)

        # ---- Alarm parameter extraction ----
        fi = ws.cell(r, fieldinfo_col).value
        params = extract_alarm_parameters(fi)
        # Default values when parameter is missing
        ws.cell(r, param1_col).value = params[0] if len(params) >= 1 else "<No value>"
        ws.cell(r, param2_col).value = params[1] if len(params) >= 2 else "<No value>"
        ws.cell(r, param3_col).value = params[2] if len(params) >= 3 else "<No value>"
        ws.cell(r, param4_col).value = params[3] if len(params) >= 4 else "<No value>"
    custom_data(data=f"Task {self.request.id}: Processing finished.")
    if os.path.exists(input_xlsx):
        os.remove(input_xlsx)
    wb.save(output_xlsx)
    os.makedirs(f"UserFiles/{str(user_id)}/alarmsTPuni", exist_ok=True)
    shutil.move(output_xlsx, f"UserFiles/{str(user_id)}/alarmsTPuni")
    custom_data(data=f"Task {self.request.id} finished at {time.ctime()}")
    return mark_safe(
        f'Current task has finished: <a href="{reverse("download", args=[user_id, output_xlsx, og_output_xlsx, "alarmsTPuni"])}">Download</a>'
    )


@shared_task(bind=True)
def proface_adress_translate(self, data, user_id=None):
    custom_data(data=f"Task {self.request.id} started at {time.ctime()}.")
    custom_data(data=f"Task {self.request.id}: Changing starting values.")
    change_start_val(int(data.get("startdbw")), int(data.get("startls")))
    int_data = data.get("intData", {})
    results = []
    results_json = {}
    custom_data(data=f"Task {self.request.id}: Reading data for PLC.")
    for i in int_data["plc"]:
        if len(int_data["plc"][i]) > 1:
            results.append(
                dbx_to_ls(
                    int(int_data["plc"][i][0].get("word")),
                    int(int_data["plc"][i][1].get("bit")),
                )
            )
        else:
            results.append(dbb_to_ls_word(int(int_data["plc"][i][0].get("word"))))
    custom_data(data=f"Task {self.request.id}: Adding data for PLC.")
    custom_data(data=f"Task {self.request.id}: Reading data for Proface.")
    for i in int_data["proface"]:
        if len(int_data["proface"][i]) > 1:
            results.append(
                ls_bit_to_dbx(
                    int(int_data["proface"][i][0].get("word")),
                    int(int_data["proface"][i][1].get("bit")),
                )
            )
        else:
            results.append(ls_to_dbb_bytes(int(int_data["proface"][i][0].get("word"))))
    custom_data(data=f"Task {self.request.id}: Adding data for Proface.")
    custom_data(data=f"Task {self.request.id}: Collecting data.")
    for i, value in enumerate(results):
        results_json[i] = value
    custom_data(data=f"Task {self.request.id} finished at {time.ctime()}")
    return json.dumps(results_json)


@shared_task(bind=True)
def merge_xlsx(
    self,
    og_xlsx,
    og_names,
    og_values,
    fix_xlsx,
    fix_names,
    fix_values,
    output_xlsx,
    user_id=None,
):
    custom_data(data=f"Task {self.request.id} started at {time.ctime()}.")
    custom_data(data=f"Task {self.request.id}: Saving file name.")
    og_output_xlsx = output_xlsx
    output_xlsx = f"{user_id}_{int(time.time())}_{output_xlsx}"
    custom_data(data=f"Task {self.request.id}: Reading Excel files.")
    df_fix = pandas.read_excel(fix_xlsx)
    df_og = pandas.read_excel(og_xlsx)
    custom_data(
        data=f"Task {self.request.id}: Changing values from alphabetical to numeric."
    )
    og_values = ord(og_values.upper()) - 65
    og_names = ord(og_names.upper()) - 65
    custom_data(data=f"Task {self.request.id}: Checking for missing columns.")
    while len(df_og.columns) <= max(og_names, og_values):
        df_og[f"Column_{df_og.columns[og_names]}"] = None

    fixes = dict(
        zip(
            df_fix.iloc[:, ord(fix_names.upper()) - 65],
            df_fix.iloc[:, ord(fix_values.upper()) - 65],
        )
    )
    custom_data(data=f"Task {self.request.id}: Starting first chunk.")
    first_chunk = True
    chunk_size = 1000
    custom_data(
        data=f"Task {self.request.id}: Starting checking columns in chunks of {chunk_size}."
    )
    for i in range(0, len(df_og), chunk_size):
        custom_data(data=f"Task {self.request.id}: {i} of {len(df_og)}.")
        df_chunk = df_og.iloc[i : i + chunk_size].copy()

        mask = df_chunk.iloc[:, og_names].isin(fixes)
        df_chunk.loc[mask, df_og.columns[og_values]] = df_chunk.loc[
            mask, df_og.columns[og_names]
        ].map(fixes)

        mode = "w" if first_chunk else "a"
        if_exists = "overlay" if not first_chunk else None
        with pandas.ExcelWriter(
            output_xlsx, engine="openpyxl", mode=mode, if_sheet_exists=if_exists
        ) as writer:
            startrow = 0 if first_chunk else writer.book.active.max_row
            df_chunk.to_excel(
                writer, index=False, header=first_chunk, startrow=startrow
            )

        first_chunk = False
    custom_data(data=f"Task {self.request.id}: Finished checking.")
    custom_data(data=f"Task {self.request.id}: Saving file.")
    os.makedirs(f"UserFiles/{str(user_id)}/FixedXLSX", exist_ok=True)
    shutil.move(output_xlsx, f"UserFiles/{str(user_id)}/FixedXLSX")
    custom_data(data=f"Task {self.request.id}: Deleting old files.")
    if os.path.exists(og_xlsx):
        os.remove(og_xlsx)
    if os.path.exists(fix_xlsx):
        os.remove(fix_xlsx)
    custom_data(data=f"Task {self.request.id} finished at {time.ctime()}")
    return mark_safe(
        f'Current task has finished: <a href="{reverse("download", args=[user_id, output_xlsx, og_output_xlsx, "FixedXLSX"])}">Download</a>'
    )
